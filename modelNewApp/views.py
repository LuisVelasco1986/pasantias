from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, user_logged_in
from django.contrib.auth.decorators import login_required
import secrets
from django.core.mail import send_mail
from django.db.models import Q, Max, Count, F, IntegerField

from django.db.models.functions import Cast, TruncDate, TruncDay, TruncWeek, TruncMonth, TruncHour, ExtractHour

import datetime
from datetime import datetime, date

from django.utils import timezone
from django.utils.timezone import now, timedelta, localdate, make_aware, localtime

from django.http import JsonResponse
from .models import Persona

from django.core.paginator import Paginator

from modelNewApp.models import *
from modelNewApp.decorators import solo_admin

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

from reportlab.pdfbase.pdfmetrics import stringWidth

from .utils.email import enviar_correo

import csv

class HomeView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect("modelNewApp:control")
        else:
            return render(request, "pages/index.html")

def privacy(request):
    return render(request, "pages/privacy.html")

def terms(request):
    return render(request, "pages/terms.html")


def login_view(request):
    # Verificación de sesión activa (corregido el nombre de la propiedad)
    if request.user.is_authenticated:
        return redirect("modelNewApp:control")

    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = authenticate(request, username=email, password=password)

        if user is not None:
            # --- NUEVA LÓGICA DE VERIFICACIÓN ---
            # Buscamos la persona asociada a este usuario
            # Nota: Ajusta 'user.persona' según como se llame la relación en tu modelo
            persona = getattr(user, 'persona', None)

            if persona and not persona.activo:
                messages.error(request, "Tu cuenta de empleado está desactivada. Contacta al administrador.")
                return redirect("modelNewApp:home")
            # ------------------------------------

            login(request, user)

            # Verificamos si debe cambiar contraseña (proceso de tu mensaje anterior)
            if getattr(user, 'debe_cambiar_password', False):
                return redirect('modelNewApp:cambiar_password')

            return redirect("modelNewApp:control")
        else:
            messages.error(request, "Correo o contraseña incorrectos.")
            return redirect("modelNewApp:home")

    return redirect("modelNewApp:home")


def logout_view(request):
    logout(request)
    return redirect('modelNewApp:home')


@login_required
def cambiar_password(request):
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "Las contraseñas no coinciden.")
        else:
            request.user.set_password(password1)
            request.user.debe_cambiar_password = False
            request.user.save()
            messages.success(request, "Contraseña actualizada correctamente.")
            return redirect('modelNewApp:home')

    return render(request, 'pages/cambiar_password.html')


@login_required
def control(request):
    if request.method == "POST":

        tipo = request.POST.get("tipo_ingreso")

        if tipo == "PIE":
            return redirect("modelNewApp:control_pie")
        elif tipo == "VEHICULO":
            return redirect("modelNewApp:control_vehiculo")

    return render(request, "pages/control.html")


from django.db import transaction

@login_required
def control_pie(request):
    departamentos = Departamento.objects.all().order_by('nombre')
    contexto = {"departamentos": departamentos}

    if request.method == "POST":
        # Capturamos si es visitante comparando con los valores posibles del checkbox
        visitante_raw = request.POST.get("visitante")
        es_visitante = visitante_raw in ["checked", "on", "true"]

        es_ingreso = "ingreso" in request.POST
        es_salida = "salida" in request.POST
        es_forzar = "forzar_salida" in request.POST

        # IMPORTANTE: Este diccionario controla lo que vuelve al formulario
        data_retorno = {
            "es_visitante": es_visitante,  # Cambiamos el nombre para evitar confusiones
            "codigo": request.POST.get("codigo"),
            "cedula_visitante": request.POST.get("cedula_visitante"),
            "nombre_visitante": request.POST.get("nombre_visitante"),
            "apellido_visitante": request.POST.get("apellido_visitante"),
            "departamento": request.POST.get("departamento"),
        }

        # --- 1. LÓGICA DE FORZAR SALIDA ---
        if es_forzar:
            motivo = request.POST.get("motivo_forzar_salida")
            persona = None
            if es_visitante:
                persona = Persona.objects.filter(cedula=request.POST.get("cedula_visitante")).first()
            else:
                persona = Persona.objects.filter(codigo_p00=request.POST.get("codigo")).first()

            if persona:
                ultimo_mov = RegistroAcceso.objects.filter(id_persona=persona).last()
                if ultimo_mov and ultimo_mov.tipo_movimiento == "INGRESO":
                    RegistroAcceso.objects.create(
                        id_persona=persona, fecha_hora=datetime.now(),
                        tipo_movimiento="EGRESO", observacion=motivo
                    )
                    contexto["mensaje"] = "Salida forzada registrada."
                else:
                    contexto.update({"mensaje_error": "La persona no está dentro.", "data": data_retorno})
            else:
                contexto.update({"mensaje_error": "Identificación no encontrada.", "data": data_retorno})

        # --- 2. LÓGICA DE REGISTRO NORMAL ---
        else:
            if es_visitante:
                cedula = request.POST.get("cedula_visitante")
                visitante = Persona.objects.filter(cedula=cedula).first()

                if visitante:
                    ultimo_mov = RegistroAcceso.objects.filter(id_persona=visitante).last()
                    ya_esta_dentro = ultimo_mov and ultimo_mov.tipo_movimiento == "INGRESO"

                    if ya_esta_dentro and es_ingreso:
                        contexto.update({"mensaje_error": "La persona ya ingresó.", "data": data_retorno})
                    elif not ya_esta_dentro and es_salida:
                        contexto.update({"mensaje_error": "No tiene entrada previa.", "data": data_retorno})
                    else:
                        visitante.nombres = request.POST.get("nombre_visitante")
                        visitante.apellidos = request.POST.get("apellido_visitante")
                        visitante.save()
                        reg = RegistroAcceso(id_persona=visitante, fecha_hora=datetime.now())
                        if es_ingreso:
                            depto = Departamento.objects.filter(nombre=request.POST.get("departamento")).first()
                            if not depto:
                                contexto.update({"mensaje_error": "Seleccione departamento.", "data": data_retorno})
                                return render(request, "pages/control_pie.html", contexto)
                            reg.tipo_movimiento, reg.departamento_destino = "INGRESO", depto
                        else:
                            reg.tipo_movimiento = "EGRESO"
                        reg.save()
                        contexto["mensaje"] = "Ingreso/salida efectuada correctamente."
                else:
                    if es_ingreso:
                        depto = Departamento.objects.filter(nombre=request.POST.get("departamento")).first()
                        if not depto:
                            contexto.update({"mensaje_error": "Seleccione departamento.", "data": data_retorno})
                            return render(request, "pages/control_pie.html", contexto)
                        try:
                            with transaction.atomic():
                                nuevo_v = Persona.objects.create(
                                    id_tipo=TipoEmpleado.objects.get(nombre_tipo="Visitante"),
                                    nombres=request.POST.get("nombre_visitante"),
                                    apellidos=request.POST.get("apellido_visitante"), cedula=cedula
                                )
                                RegistroAcceso.objects.create(
                                    id_persona=nuevo_v, tipo_movimiento="INGRESO",
                                    departamento_destino=depto, fecha_hora=datetime.now()
                                )
                                contexto["mensaje"] = "Visitante creado e ingresado."
                        except Exception as e:
                            contexto.update({"mensaje_error": f"Error: {e}", "data": data_retorno})
                    else:
                        contexto.update({"mensaje_error": "Visitante no existe.", "data": data_retorno})
            else:
                # EMPLEADOS
                codigo = request.POST.get("codigo")
                empleado = Persona.objects.filter(codigo_p00=codigo).first()
                if empleado:
                    ultimo_mov = RegistroAcceso.objects.filter(id_persona=empleado).last()
                    ya_esta_dentro = ultimo_mov and ultimo_mov.tipo_movimiento == "INGRESO"
                    if ya_esta_dentro and es_ingreso:
                        contexto.update({"mensaje_error": "Empleado ya ingresó.", "data": data_retorno})
                    elif not ya_esta_dentro and es_salida:
                        contexto.update({"mensaje_error": "Empleado no tiene entrada.", "data": data_retorno})
                    else:
                        reg = RegistroAcceso(id_persona=empleado, fecha_hora=datetime.now(),
                                             tipo_movimiento="INGRESO" if es_ingreso else "EGRESO")
                        if es_ingreso: reg.departamento_destino = empleado.departamento
                        reg.save()
                        contexto["mensaje"] = "Ingreso/salida efectuada correctamente."
                else:
                    contexto.update({"mensaje_error": "Código P00 no existe.", "data": data_retorno})

    return render(request, "pages/control_pie.html", contexto)

def obtener_persona(request):
    if request.POST.get("visitante"):
        return Persona.objects.filter(
            cedula=request.POST.get("cedula_visitante")
        ).first()
    else:
        return Persona.objects.filter(
            codigo_p00=request.POST.get("codigo")
        ).first()

def obtener_o_crear_vehiculo(request, marca, modelo):
    placa = request.POST.get("placa")
    codigo = request.POST.get("codigo_vehiculo")

    vehiculo = Vehiculo.objects.filter(placa=placa).first()

    if vehiculo:
        vehiculo.marca = marca
        vehiculo.modelo = modelo
        if codigo:
            vehiculo.codigo = codigo
        vehiculo.save()
    else:
        vehiculo = Vehiculo.objects.create(
            placa=placa,
            marca=marca,
            modelo=modelo,
            codigo=codigo if codigo else ""
        )

    return vehiculo

def ultimo_movimiento_persona(persona):
    return RegistroAcceso.objects.filter(
        id_persona=persona
    ).order_by("-fecha_hora").first()

def ultimo_movimiento_vehiculo(vehiculo):
    return RegistroAcceso.objects.filter(
        vehiculo=vehiculo
    ).order_by("-fecha_hora").first()

@login_required
def control_vehiculo(request):
    # Cargamos catálogos para el contexto
    context = {
        "departamentos": Departamento.objects.all().order_by('nombre'),
        "marcas": Marca.objects.all().order_by('nombre'),
        "modelos": Modelo.objects.all().order_by('nombre')
    }

    if request.method == "POST":
        # Capturamos el booleano correctamente
        visitante_raw = request.POST.get("visitante")
        es_visitante = visitante_raw in ["checked", "on", "true"]

        # Creamos el diccionario data para devolverlo en caso de error
        data_retorno = {
            "es_visitante": es_visitante,
            "cedula_visitante": request.POST.get("cedula_visitante"),
            "nombre_visitante": request.POST.get("nombre_visitante"),
            "apellido_visitante": request.POST.get("apellido_visitante"),
            "placa": request.POST.get("placa"),
            "marca": request.POST.get("marca"),
            "modelo": request.POST.get("modelo"),
            "codigo_vehiculo": request.POST.get("codigo_vehiculo"),
            "codigo": request.POST.get("codigo"), # Código P00
            "departamento": request.POST.get("departamento"),
        }

        try:
            with transaction.atomic():
                # Captura de datos básicos usando el booleano normalizado
                marca_txt = request.POST.get("marca", "").strip()
                modelo_txt = request.POST.get("modelo", "").strip()

                marca, _ = Marca.objects.get_or_create(
                    nombre__iexact=marca_txt, defaults={"nombre": marca_txt}
                )
                modelo, _ = Modelo.objects.get_or_create(
                    nombre__iexact=modelo_txt, marca=marca, defaults={"nombre": modelo_txt, "marca": marca}
                )

                vehiculo = obtener_o_crear_vehiculo(request, marca, modelo)
                persona = obtener_persona(request)

                if not persona and es_visitante:
                    persona = Persona.objects.create(
                        id_tipo=TipoEmpleado.objects.get(nombre_tipo="Visitante"),
                        nombres=request.POST.get("nombre_visitante"),
                        apellidos=request.POST.get("apellido_visitante"),
                        cedula=request.POST.get("cedula_visitante")
                    )

                if not persona:
                    context["mensaje_error"] = "No existe una persona con esos datos."
                    context["data"] = data_retorno # Devolvemos los datos limpios
                    raise Exception("Rollback: Persona no encontrada")

                # Verificación de estados
                ultimo_p = ultimo_movimiento_persona(persona)
                ultimo_v = ultimo_movimiento_vehiculo(vehiculo)
                persona_dentro = ultimo_p and ultimo_p.tipo_movimiento == "INGRESO"
                vehiculo_dentro = ultimo_v and ultimo_v.tipo_movimiento == "INGRESO"

                # Lógica de Movimientos
                if "salida" in request.POST or "forzar_salida" in request.POST:
                    if not persona_dentro or not vehiculo_dentro:
                        if not persona_dentro:
                            context["mensaje_error"] = "No se puede registrar salida: La persona no se encuentra dentro."
                        if not vehiculo_dentro:
                            context["mensaje_error"] = "No se puede registrar salida: El vehículo no se encuentra dentro."
                        context["data"] = data_retorno
                        raise Exception("Rollback: Estado inválido para salida")

                    RegistroAcceso.objects.create(
                        id_persona=persona, vehiculo=vehiculo, tipo_movimiento="EGRESO",
                        fecha_hora=timezone.now(),
                        observacion=request.POST.get("motivo_forzar_salida") if "forzar_salida" in request.POST else ""
                    )
                    if es_visitante:
                        visitante = Persona.objects.filter(cedula=request.POST.get("cedula_visitante")).first()
                        visitante.nombres = request.POST.get("nombre_visitante")
                        visitante.apellidos = request.POST.get("apellido_visitante")
                        visitante.save()
                    context["mensaje"] = "Salida registrada correctamente."

                elif "ingreso" in request.POST:
                    if persona_dentro or vehiculo_dentro:
                        if persona_dentro:
                            context["mensaje_error"] = "No se puede registrar ingreso: La persona se encuentra dentro."
                        if vehiculo_dentro:
                            context["mensaje_error"] = "No se puede registrar ingreso: El vehículo se encuentra dentro."
                        context["data"] = data_retorno
                        raise Exception("Rollback: Ya están dentro")

                    depto = None
                    if es_visitante:
                        depto_nombre = request.POST.get("departamento")
                        if not depto_nombre:
                            context["mensaje_error"] = "Debe seleccionar un departamento."
                            context["data"] = data_retorno
                            raise Exception("Rollback: Falta departamento")
                        depto = Departamento.objects.filter(nombre=depto_nombre).first()
                        visitante = Persona.objects.filter(cedula=request.POST.get("cedula_visitante")).first()
                        visitante.nombres = request.POST.get("nombre_visitante")
                        visitante.apellidos = request.POST.get("apellido_visitante")
                        visitante.save()
                    else:
                        depto = persona.departamento

                    RegistroAcceso.objects.create(
                        id_persona=persona, vehiculo=vehiculo, tipo_movimiento="INGRESO",
                        fecha_hora=timezone.now(), departamento_destino=depto
                    )

                    context["mensaje"] = "Entrada registrada correctamente."

        except Exception as e:
            print(f"Transacción revertida: {e}")
            if "mensaje_error" not in context:
                context["mensaje_error"] = "Ocurrió un error al procesar la solicitud."
            # Aseguramos que data siempre viaje en el catch si no se definió antes
            if "data" not in context:
                context["data"] = data_retorno

    return render(request, "pages/control_vehiculo.html", context)

def buscar_persona_por_cedula(request):
    cedula = request.GET.get("cedula")

    if not cedula:
        return JsonResponse({"found": False})

    persona = Persona.objects.filter(cedula=cedula).first()

    if persona:
        return JsonResponse({
            "found": True,
            "nombres": persona.nombres,
            "apellidos": persona.apellidos
        })

    return JsonResponse({"found": False})


def buscar_vehiculo_por_placa(request):
    placa = request.GET.get("placa")

    try:
        vehiculo = Vehiculo.objects.get(placa__iexact=placa)
        return JsonResponse({
            "existe": True,
            "marca": vehiculo.marca.nombre,
            "modelo": vehiculo.modelo.nombre,
            "codigo": vehiculo.codigo,
        })
    except Vehiculo.DoesNotExist:
        return JsonResponse({"existe": False})


@solo_admin
@login_required
def dashboard(request):
    today = timezone.localdate()

    # ─────────────────────────────────────────────
    # Gráfica: Ingresos últimos 7 días
    # ─────────────────────────────────────────────
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]

    ingresos_por_dia = []
    for day in last_7_days:
        count = RegistroAcceso.objects.filter(
            tipo_movimiento='INGRESO',
            fecha_hora__date=day
        ).count()

        ingresos_por_dia.append({
            'fecha': day.strftime('%d/%m'),
            'cantidad': count
        })

    # ─────────────────────────────────────────────
    # Personas dentro del edificio
    # (último movimiento = INGRESO)
    # ─────────────────────────────────────────────
    ultimos_movimientos = (
        RegistroAcceso.objects
        .values('id_persona')
        .annotate(ultima_fecha=Max('fecha_hora'))
    )

    dentro_ids = []

    for mov in ultimos_movimientos:
        ultimo = RegistroAcceso.objects.get(
            id_persona_id=mov['id_persona'],
            fecha_hora=mov['ultima_fecha']
        )
        if ultimo.tipo_movimiento == 'INGRESO':
            dentro_ids.append(ultimo.id)

    dentro = RegistroAcceso.objects.filter(id__in=dentro_ids).order_by("-fecha_hora")

    # ─────────────────────────────────────────────
    # Ingresos y salidas de HOY
    # ─────────────────────────────────────────────
    ingresos = RegistroAcceso.objects.filter(
        tipo_movimiento='INGRESO',
        fecha_hora__date=today
    )

    salidas = RegistroAcceso.objects.filter(
        tipo_movimiento='EGRESO',
        fecha_hora__date=today
    )

    # ─────────────────────────────────────────────
    # Vehículos dentro del edificio
    # (último movimiento = INGRESO)
    # ─────────────────────────────────────────────
    ultimos_movimientos_vehiculos = (
        RegistroAcceso.objects
        .filter(vehiculo__isnull=False)
        .values('vehiculo')
        .annotate(ultima_fecha=Max('fecha_hora'))
    )

    vehiculos_dentro_ids = []

    for mov in ultimos_movimientos_vehiculos:
        ultimo = RegistroAcceso.objects.get(
            vehiculo_id=mov['vehiculo'],
            fecha_hora=mov['ultima_fecha']
        )

        if ultimo.tipo_movimiento == 'INGRESO':
            vehiculos_dentro_ids.append(ultimo.id)

    vehiculos_dentro = RegistroAcceso.objects.filter(
        id__in=vehiculos_dentro_ids
    ).order_by("-fecha_hora")

    # ─────────────────────────────────────────────
    # Total personas
    # ─────────────────────────────────────────────
    total = Persona.objects.count()

    context = {
        "dentro": dentro,
        "total": total,
        "ingresos": ingresos,
        "salidas": salidas,
        "ingresos_por_dia": ingresos_por_dia,
        "vehiculos_dentro": vehiculos_dentro,
    }

    return render(request, "pages/dashboard_main.html", context)


@solo_admin
@login_required
def empleados(request):
    empleados = Persona.objects.all().order_by("nombres")
    tipos = TipoEmpleado.objects.all().order_by("nombre_tipo")

    search = request.GET.get('search', '')
    estado = request.GET.get('estado')
    tipo_id = request.GET.get('tipo')
    orden = request.GET.get('orden')

    if search:
        empleados = empleados.filter(
            Q(nombres__icontains=search) |
            Q(apellidos__icontains=search) |
            Q(codigo_p00__icontains=search)
        )

    tipo_seleccionado = None
    if tipo_id:
        tipo_seleccionado = TipoEmpleado.objects.get(id=tipo_id).nombre_tipo

    if estado == 'activo':
        empleados = empleados.filter(activo=True)
    elif estado == 'inactivo':
        empleados = empleados.filter(activo=False)

    if tipo_id:
        empleados = empleados.filter(id_tipo_id=tipo_id)

    if orden == "nombre":
        empleados = empleados.order_by("nombres", "apellidos")
    elif orden == "-nombre":
        empleados = empleados.order_by("-nombres", "-apellidos")
    elif orden == "fecha":
        empleados = empleados.order_by("fecha_registro")
    elif orden == "-fecha":
        empleados = empleados.order_by("-fecha_registro")
    elif orden == "codigo":
        empleados = empleados.order_by("codigo_p00")
    elif orden == "-codigo":
        empleados = empleados.order_by("-codigo_p00")
    elif orden == "-estado":
        empleados = empleados.order_by("-activo")
    elif orden == "estado":
        empleados = empleados.order_by("activo")
    else:
        empleados = empleados.order_by("nombres")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'partials/personas_list.html', {
            'empleados': empleados,
            "tipos": tipos,
            "tipo_seleccionado": tipo_seleccionado,
            "orden_actual": orden
        })

    dict = {"empleados": empleados, "tipos": tipos, "tipo_seleccionado": tipo_seleccionado, "orden_actual": orden}

    return render(request, "pages/empleados.html", dict)

@solo_admin
@login_required
def empleados_agregar(request):
    if request.POST:
        sexo = request.POST.get('sexo')
        tipos = TipoEmpleado.objects.all().order_by('nombre_tipo').exclude(nombre_tipo="Visitante")
        roles_ids = request.POST.getlist('roles')
        es_admin = Rol.objects.filter(
            id__in=roles_ids,
            nombre_rol__iexact="Administrador"
        ).exists()
        departamentos = Departamento.objects.all().order_by("nombre")

        if not sexo:
            messages.error(request, "Debe seleccionar un sexo.")
            roles = Rol.objects.all().order_by('nombre_rol')
            dict = {"error": "Debe seleccionar un sexo", "data": request.POST, "tipos": tipos, "roles": roles,
                    "departamentos": departamentos}
            return render(request, "pages/empleados_agregar.html", dict)
        else:
            if request.POST.get("administrador"):
                if roles_ids:

                    empleado = Persona()
                    empleado.nombres = request.POST.get('nombres')
                    empleado.apellidos = request.POST.get('apellidos')
                    if Persona.objects.filter(cedula=request.POST.get('cedula')).exists():
                        messages.error(request, "Ya existe un empleado con esa cédula.")
                        return render(
                            request,
                            "pages/empleados_agregar.html",
                            {
                                "error": "Ya existe un empleado con esa cédula.",
                                "data": request.POST,
                                "roles": Rol.objects.all(),
                                "tipos": TipoEmpleado.objects.all().exclude(nombre_tipo="Visitante"),
                                "departamentos": departamentos,
                            })
                    else:
                        empleado.cedula = request.POST.get('cedula')
                    if Persona.objects.filter(codigo_p00=request.POST.get('codigo')).exists():
                        messages.error(request, "Ya existe un empleado con ése código P00.")
                        return render(
                            request,
                            "pages/empleados_agregar.html",
                            {
                                "error": "Ya existe un empleado con ése código P00.",
                                "data": request.POST,
                                "roles": Rol.objects.all(),
                                "tipos": TipoEmpleado.objects.all().exclude(nombre_tipo="Visitante"),
                                "departamentos": departamentos,
                            })
                    else:
                        empleado.codigo_p00 = request.POST.get('codigo')
                    if request.POST.get('departamento'):
                        empleado.departamento = Departamento.objects.filter(id=request.POST.get('departamento')).first()
                    else:
                        empleado.departamento = None
                    empleado.sexo = request.POST.get('sexo')
                    empleado.id_tipo = TipoEmpleado.objects.get(id=request.POST.get('tipo_persona'))
                    if 'foto_perfil' in request.FILES:
                        empleado.foto_perfil = request.FILES['foto_perfil']
                    if Usuario.objects.filter(email=request.POST.get('email')).exists():
                        messages.error(request, "Ya existe un usuario con ése email.")
                        return render(
                            request,
                            "pages/empleados_agregar.html",
                            {
                                "error": "Ya existe un usuario con ése email.",
                                "data": request.POST,
                                "roles": Rol.objects.all(),
                                "tipos": TipoEmpleado.objects.all().exclude(nombre_tipo="Visitante"),
                                "departamentos": departamentos,
                            })
                    empleado.save()

                    usuario = Usuario()
                    usuario.email = request.POST.get('email')
                    usuario.username = request.POST.get('email')
                    usuario.first_name = request.POST.get('nombres')
                    usuario.last_name = request.POST.get('apellidos')
                    usuario.persona = empleado
                    password_temporal = secrets.token_urlsafe(8)
                    usuario.set_password(password_temporal)
                    usuario.debe_cambiar_password = True
                    # 👉 permisos admin
                    usuario.is_staff = True  # puede entrar al admin
                    usuario.is_active = True  # usuario activo
                    usuario.is_superuser = es_admin
                    usuario.save()
                    usuario.roles.set(roles_ids)

                    html_content = f"""
                    <h3>Hola {usuario.first_name},</h3>
                    <p>Se ha creado una cuenta para ti.</p>
                    <p><strong>Usuario:</strong> {usuario.email}</p>
                    <p><strong>Contraseña temporal:</strong> {password_temporal}</p>
                    <p>Por seguridad deberás cambiar esta contraseña al iniciar sesión.</p>
                    """

                    enviar_correo(
                        usuario.email,
                        "Acceso al sistema",
                        html_content
                    )

                else:
                    messages.error(request, "Debe seleccionar al menos un rol.")
                    roles = Rol.objects.all().order_by('nombre_rol')
                    dict = {"error": "Debe seleccionar al menos un rol.", "data": request.POST, "tipos": tipos,
                            "roles": roles, "departamentos": departamentos}
                    return render(request, "pages/empleados_agregar.html", dict)
            else:
                empleado = Persona()
                empleado.nombres = request.POST.get('nombres')
                empleado.apellidos = request.POST.get('apellidos')
                empleado.cedula = request.POST.get('cedula')
                empleado.codigo_p00 = request.POST.get('codigo')
                if request.POST.get('departamento'):
                    empleado.departamento = Departamento.objects.filter(id=request.POST.get('departamento')).first()
                else:
                    empleado.departamento = None
                empleado.sexo = request.POST.get('sexo')
                empleado.id_tipo = TipoEmpleado.objects.get(id=request.POST.get('tipo_persona'))
                if 'foto_perfil' in request.FILES:
                    empleado.foto_perfil = request.FILES['foto_perfil']
                empleado.save()

            return redirect("modelNewApp:empleados")

    tipos = TipoEmpleado.objects.all().order_by('nombre_tipo').exclude(nombre_tipo="Visitante")
    roles = Rol.objects.all().order_by('nombre_rol')
    departamentos = Departamento.objects.all().order_by("nombre")
    dict = {"tipos": tipos, "roles": roles, "departamentos": departamentos}
    return render(request, "pages/empleados_agregar.html", dict)


@solo_admin
@login_required
def empleados_detalles(request, id):
    persona = get_object_or_404(Persona, id=id)

    historial_accesos = (
        RegistroAcceso.objects
        .filter(id_persona=persona)
        .select_related('departamento_destino')
        .order_by('-fecha_hora')
    )

    # ──────────────── Paginación ────────────────
    paginator = Paginator(historial_accesos, 20)  # 10 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    dict = {"persona": persona, "page_obj": page_obj, }
    return render(request, "pages/empleados_detalles.html", dict)


@solo_admin
@login_required
def empleados_eliminar(request, id):
    persona = get_object_or_404(Persona, id=id)

    if hasattr(persona, 'empleado') and persona.empleado:
        persona.empleado.delete()

    persona.delete()
    return redirect("modelNewApp:empleados")


@solo_admin
@login_required
def empleados_editar(request, id):
    persona = get_object_or_404(Persona, id=id)

    roles_persona = []
    if persona and hasattr(persona, 'empleado'):
        usuario = persona.empleado
        roles_persona = usuario.roles.values_list('id', flat=True)

    if request.POST:
        sexo = request.POST.get('sexo')
        tipos = TipoEmpleado.objects.all().order_by('nombre_tipo')
        roles_ids = request.POST.getlist('roles')
        es_admin = Rol.objects.filter(
            id__in=roles_ids,
            nombre_rol__iexact="Administrador"
        ).exists()
        departamentos = Departamento.objects.all().order_by("nombre")

        persona = get_object_or_404(Persona, id=id)

        roles_persona = []
        usuario = False
        if persona and hasattr(persona, 'empleado'):
            usuario = persona.empleado
            roles_persona = usuario.roles.values_list('id', flat=True)

        if not sexo:
            messages.error(request, "Debe seleccionar un sexo.")
            dict = {"error": "Debe seleccionar un sexo", "data": request.POST, "tipos": tipos, "persona": persona,
                    "departamentos": departamentos}
            return render(request, "pages/empleados_editar.html", dict)
        else:
            persona.nombres = request.POST.get('nombres')
            persona.apellidos = request.POST.get('apellidos')
            if Persona.objects.filter(cedula=request.POST.get('cedula')).exclude(id=persona.id).exists():
                messages.error(request, "Ya existe un empleado con esa cédula.")
                return render(
                    request,
                    "pages/empleados_editar.html",
                    {
                        "error": "Ya existe un empleado con esa cédula.",
                        "data": request.POST,
                        "roles": Rol.objects.all(),
                        "tipos": TipoEmpleado.objects.all().order_by('nombre_tipo'),
                        "persona": persona,
                        "roles_persona": roles_persona,
                        "departamentos": departamentos,
                    })
            else:
                persona.cedula = request.POST.get('cedula')
            if Persona.objects.filter(codigo_p00=request.POST.get('codigo')).exclude(id=persona.id).exists():
                messages.error(request, "Ya existe un empleado con ése código P00.")
                return render(
                    request,
                    "pages/empleados_editar.html",
                    {
                        "error": "Ya existe un empleado con ése código P00.",
                        "data": request.POST,
                        "roles": Rol.objects.all(),
                        "tipos": TipoEmpleado.objects.all().order_by('nombre_tipo'),
                        "persona": persona,
                        "roles_persona": roles_persona,
                        "departamentos": departamentos,
                    })
            else:
                persona.codigo_p00 = request.POST.get('codigo')
            if request.POST.get('departamento'):
                persona.departamento = Departamento.objects.filter(id=request.POST.get('departamento')).first()
            else:
                persona.departamento = None
            persona.sexo = request.POST.get('sexo')
            persona.id_tipo = TipoEmpleado.objects.get(id=request.POST.get('tipo_persona'))
            if 'foto_perfil' in request.FILES:
                persona.foto_perfil = request.FILES['foto_perfil']
            if Usuario.objects.filter(email=request.POST.get('email')).exists() and Usuario.objects.filter(email=request.POST.get('email')).first() != persona.empleado:
                messages.error(request, "Ya existe un usuario con ése email.")
                return render(
                    request,
                    "pages/empleados_editar.html",
                    {
                        "error": "Ya existe un usuario con ése email.",
                        "data": request.POST,
                        "roles": Rol.objects.all(),
                        "tipos": TipoEmpleado.objects.all().order_by('nombre_tipo'),
                        "persona": persona,
                        "roles_persona": roles_persona,
                        "departamentos": departamentos,
                    })
            persona.save()

            if request.POST.get("administrador"):
                if roles_ids:

                    if usuario:

                        usuario.username = request.POST.get('email')
                        usuario.first_name = request.POST.get('nombres')
                        usuario.last_name = request.POST.get('apellidos')
                        usuario.persona = persona
                        # 👉 permisos admin
                        usuario.is_staff = True  # puede entrar al admin
                        usuario.is_active = True  # usuario activo
                        usuario.is_superuser = es_admin
                        usuario.save()
                        usuario.roles.set(roles_ids)

                        if request.POST.get('email') != usuario.email:
                            password_temporal = secrets.token_urlsafe(8)
                            usuario.set_password(password_temporal)
                            usuario.debe_cambiar_password = True
                            usuario.email = request.POST.get('email')
                            usuario.username = request.POST.get('email')
                            usuario.save()

                            html_content = f"""
                                    Hola {usuario.first_name},

                                    Se ha cambiado su dirección de correo electrónico.

                                    Usuario: {usuario.email}
                                    Contraseña temporal: {password_temporal}

                                    Por seguridad deberás cambiar esta contraseña al iniciar sesión.
                                    """

                            enviar_correo(
                                usuario.email,
                                "Acceso al sistema",
                                html_content
                            )
                    else:
                        usuario = Usuario()
                        usuario.email = request.POST.get('email')
                        usuario.username = request.POST.get('email')
                        usuario.first_name = request.POST.get('nombres')
                        usuario.last_name = request.POST.get('apellidos')
                        usuario.persona = persona
                        password_temporal = secrets.token_urlsafe(8)
                        usuario.set_password(password_temporal)
                        usuario.debe_cambiar_password = True
                        # 👉 permisos admin
                        usuario.is_staff = True  # puede entrar al admin
                        usuario.is_active = True  # usuario activo
                        usuario.is_superuser = es_admin
                        usuario.save()
                        usuario.roles.set(roles_ids)

                        html_content = f"""
                                        Hola {usuario.first_name},

                                        Se ha creado una cuenta para ti.
        
                                        Usuario: {usuario.email}
                                        Contraseña temporal: {password_temporal}
        
                                        Por seguridad deberás cambiar esta contraseña al iniciar sesión.
                                        """

                        enviar_correo(
                            usuario.email,
                            "Acceso al sistema",
                            html_content
                        )

                else:
                    messages.error(request, "Debe seleccionar al menos un rol.")
                    roles = Rol.objects.all().order_by('nombre_rol')
                    dict = {"error": "Debe seleccionar al menos un rol.", "data": request.POST, "tipos": tipos,
                            "roles": roles, "persona": persona, "departamentos": departamentos}
                    return render(request, "pages/empleados_editar.html", dict)

                if request.user.persona:
                    if request.user.persona.id == id:
                        return redirect("modelNewApp:perfil")
                    else:
                        return redirect("modelNewApp:empleados_detalles", persona.id)
                else:
                    return redirect("modelNewApp:empleados_detalles", persona.id)
            else:

                if usuario:
                    persona.empleado.delete()
                if request.user.persona:
                    if request.user.persona.id == id:
                        return redirect("modelNewApp:perfil")
                    else:
                        return redirect("modelNewApp:empleados_detalles", persona.id)
                else:
                    return redirect("modelNewApp:empleados_detalles", persona.id)

    tipos = TipoEmpleado.objects.all().order_by('nombre_tipo')
    roles = Rol.objects.all().order_by('nombre_rol')
    departamentos = Departamento.objects.all().order_by("nombre")
    dict = {"tipos": tipos, "persona": persona, "roles": roles, "roles_persona": roles_persona,
            "departamentos": departamentos}
    return render(request, "pages/empleados_editar.html", dict)


@solo_admin
@login_required
def empleados_activar(request, id):
    persona = get_object_or_404(Persona, id=id)
    persona.activo = True
    persona.save()

    return redirect("modelNewApp:empleados_detalles", persona.id)


@solo_admin
@login_required
def empleados_desactivar(request, id):
    persona = get_object_or_404(Persona, id=id)
    persona.activo = False
    persona.save()

    return redirect("modelNewApp:empleados_detalles", persona.id)


@login_required
def dashboard_control(request):
    return render(request, "pages/dashboard_control.html")


@solo_admin
@login_required
def vehiculos(request):
    vehiculos = Vehiculo.objects.all().order_by("placa")
    marcas = Marca.objects.all().order_by("nombre")
    modelos = Modelo.objects.all().order_by("nombre")

    search = request.GET.get('search', '')
    marca = request.GET.get('marca')
    modelo = request.GET.get('modelo')
    orden = request.GET.get('orden')

    if marca:
        vehiculos = vehiculos.filter(marca__id__icontains=marca)

    if modelo:
        vehiculos = vehiculos.filter(modelo__id__icontains=modelo)

    if search:
        vehiculos = vehiculos.filter(
            Q(placa__icontains=search) |
            Q(codigo__icontains=search) |
            Q(marca__nombre__icontains=search) |
            Q(modelo__nombre__icontains=search)
        )

    if orden == "placa":
        vehiculos = vehiculos.order_by("placa")
    elif orden == "-placa":
        vehiculos = vehiculos.order_by("-placa")
    elif orden == "marca":
        vehiculos = vehiculos.order_by("marca__nombre")
    elif orden == "-marca":
        vehiculos = vehiculos.order_by("-marca__nombre")
    elif orden == "modelo":
        vehiculos = vehiculos.order_by("modelo__nombre")
    elif orden == "-modelo":
        vehiculos = vehiculos.order_by("-modelo__nombre")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'partials/vehiculos_list.html', {
            'vehiculos': vehiculos,
            "orden_actual": orden
        })

    dict = {"vehiculos": vehiculos, "marcas": marcas, "modelos": modelos, "orden_actual": orden}

    return render(request, "pages/vehiculos.html", dict)

@solo_admin
@login_required
def vehiculos_agregar(request):

    if request.POST:

        placa = request.POST.get("placa")
        marca = request.POST.get("marca")
        modelo = request.POST.get("modelo")

        vehiculo = Vehiculo.objects.filter(placa=placa).first()

        if vehiculo:
            print("Ya existe esa placa")
        else:
            vehiculo = Vehiculo()
            vehiculo.placa = placa
            marca, _ = Marca.objects.get_or_create(nombre__iexact=marca, defaults={"nombre": marca})
            vehiculo.marca = marca
            modelo, _ = Modelo.objects.get_or_create(nombre__iexact=modelo, defaults={"nombre": modelo, "marca": marca})
            vehiculo.modelo = modelo
            if request.POST.get("codigo"):
                vehiculo.codigo = request.POST.get("codigo")
            vehiculo.save()

            return redirect("modelNewApp:vehiculos")

    marcas = Marca.objects.all().order_by("nombre")
    modelos = Modelo.objects.all().order_by("nombre")

    dict = {"marcas": marcas, "modelos": modelos}
    return render(request, "pages/vehiculos_agregar.html", dict)

@solo_admin
@login_required
def vehiculos_detalles(request, id):

    vehiculo = Vehiculo.objects.get(id=id)

    historial_accesos = (
        RegistroAcceso.objects
        .filter(vehiculo=vehiculo)
        .select_related('departamento_destino')
        .order_by('-fecha_hora')
    )

    # ──────────────── Paginación ────────────────
    paginator = Paginator(historial_accesos, 20)  # 10 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    dict = {"vehiculo": vehiculo, "page_obj": page_obj}

    return render(request, "pages/vehiculos_detalles.html", dict)

@solo_admin
@login_required
def vehiculos_eliminar(request, id):

    vehiculo = Vehiculo.objects.filter(id=id).first()

    if vehiculo:
        vehiculo.delete()

    return redirect("modelNewApp:vehiculos")

@solo_admin
@login_required
def vehiculos_editar(request, id):

    vehiculo = Vehiculo.objects.filter(id=id).first()
    marcas = Marca.objects.all().order_by("nombre")
    modelos = Modelo.objects.all().order_by("nombre")

    if request.POST:

        placa = request.POST.get("placa")
        marca = request.POST.get("marca")
        modelo = request.POST.get("modelo")

        vehiculo.placa = placa
        marca, _ = Marca.objects.get_or_create(nombre__iexact=marca, defaults={"nombre": marca})
        vehiculo.marca = marca
        modelo, _ = Modelo.objects.get_or_create(nombre__iexact=modelo, defaults={"nombre": modelo, "marca": marca})
        vehiculo.modelo = modelo
        if request.POST.get("codigo"):
            vehiculo.codigo = request.POST.get("codigo")
        else:
            vehiculo.codigo = ""
        vehiculo.save()

        return redirect("modelNewApp:vehiculos_detalles", vehiculo.id)

    dict = {"vehiculo": vehiculo, "marcas": marcas, "modelos": modelos}

    return render(request, "pages/vehiculos_editar.html", dict)

@solo_admin
@login_required
def estadisticos(request):
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    tipo = request.GET.get("tipo")
    tipos = TipoEmpleado.objects.all().order_by('nombre_tipo')

    registros = RegistroAcceso.objects.filter(tipo_movimiento="INGRESO")

    # -------------------------
    # Filtros por fecha
    # -------------------------
    if desde:
        registros = registros.filter(fecha_hora__date__gte=desde)

    if hasta:
        registros = registros.filter(fecha_hora__date__lte=hasta)

    # -------------------------
    # Filtro por tipo de persona
    # -------------------------
    if tipo:
        registros = registros.filter(id_persona__id_tipo__nombre_tipo=tipo)

    # -------------------------
    # TOTAL INGRESOS
    # -------------------------
    total_ingresos = registros.count()

    # -------------------------
    # PROMEDIO DIARIO (DÍAS CON ACTIVIDAD)
    # -------------------------
    dias_con_ingresos = (
        registros
        .dates("fecha_hora", "day")
        .count()
    )

    promedio_dias_activos = round(
        total_ingresos / dias_con_ingresos, 2
    ) if dias_con_ingresos > 0 else 0

    # -------------------------
    # PROMEDIO DIARIO (PERÍODO COMPLETO)
    # -------------------------
    if desde and hasta:
        dias_periodo = (
                               date.fromisoformat(hasta) -
                               date.fromisoformat(desde)
                       ).days + 1
        rango_dias = (date.fromisoformat(hasta) - date.fromisoformat(desde)).days + 1
    else:
        primer_registro = registros.order_by("fecha_hora").first()

        if primer_registro:
            fecha_inicio = localtime(primer_registro.fecha_hora).date()
            fecha_fin = localtime(timezone.now()).date()
            dias_periodo = (fecha_fin - fecha_inicio).days + 1
            rango_dias = dias_periodo
        else:
            dias_periodo = 0
            rango_dias = 0

    promedio_periodo = round(
        total_ingresos / dias_periodo, 2
    ) if dias_periodo > 0 else 0
    # ---------------------------------
    #     Nuevo dia con más ingresos
    # ---------------------------------

    dia_pico_data = (
        registros
        .annotate(dia=TruncDate('fecha_hora'))
        .values('dia')
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
        .first()
    )

    if dia_pico_data:
        dia_mas_ingresos = dia_pico_data['dia'].strftime('%d/%m/%Y')
        cantidad = dia_pico_data['cantidad']

        if cantidad >= 1000:
            cantidad_mas_ingresos_fmt = (
                f"{cantidad / 1000:.1f}k"
            ).replace(".0k", "k")
        else:
            cantidad_mas_ingresos_fmt = str(cantidad)
    else:
        dia_mas_ingresos = '-'
        cantidad_mas_ingresos_fmt = '0'

    # ----------------------------------------------------------------------------
    #     Nueva hora pico
    # ----------------------------------------------------------------------------
    datos_hora = (
        registros
        .annotate(hora_num=ExtractHour('fecha_hora'))
        .values('hora_num')
        .annotate(cantidad=Count('id'))
    )

    hora_pico_data = datos_hora.order_by('-cantidad').first()

    if hora_pico_data:
        hora_pico = datetime.strptime(
            str(hora_pico_data['hora_num']), "%H"
        ).strftime('%I:%M %p')

        cantidad = hora_pico_data['cantidad']
        cantidad_hora_pico = (
            f"{round(cantidad / 1000)}k"
            if cantidad >= 1000
            else str(cantidad)
        )
    else:
        hora_pico = "-"
        cantidad_hora_pico = "0"

    fecha_inicio = datetime.strptime(desde, '%Y-%m-%d') if desde else None
    fecha_fin = datetime.strptime(hasta, '%Y-%m-%d') if hasta else None

    # Elegir la granularidad
    if rango_dias <= 30:
        trunc = TruncDay('fecha_hora')
    elif rango_dias <= 90:
        trunc = TruncWeek('fecha_hora', week_start=1)  # semana inicia lunes
    else:
        trunc = TruncMonth('fecha_hora')

    ingresos_por_periodo = (
        registros
        .annotate(periodo=trunc)
        .values('periodo')
        .annotate(cantidad=Count('id'))
        .order_by('periodo')
    )

    for r in ingresos_por_periodo:
        if isinstance(trunc, TruncDay):
            r['periodo'] = r['periodo'].strftime('%d/%m')
        elif isinstance(trunc, TruncWeek):
            # semana: mostrar rango de la semana
            inicio_semana = r['periodo']
            fin_semana = inicio_semana + timedelta(days=6)
            r['periodo'] = f"{inicio_semana.strftime('%d/%m')} - {fin_semana.strftime('%d/%m')}"
        else:  # TruncMonth
            r['periodo'] = r['periodo'].strftime('%b %Y')  # ej: Ene 2026

    ingresos_por_hora = list(
        datos_hora.order_by('hora_num')
    )

    for r in ingresos_por_hora:
        r['hora_label'] = datetime.strptime(
            str(r['hora_num']), "%H"
        ).strftime("%I %p")

    # Tipo de persona
    ingresos_por_tipo = (
        registros
        .values(nombre=F('id_persona__id_tipo__nombre_tipo'))
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
    )

    # Tabla de accesos
    detalle_accesos = (
        registros
        .values(
            'id_persona',
            'id_persona__nombres',
            'id_persona__apellidos',
            'id_persona__id_tipo__nombre_tipo'
        )
        .annotate(
            total_ingresos=Count('id'),
            ultimo_ingreso=Max('fecha_hora')
        )
        .order_by('-total_ingresos', '-ultimo_ingreso')
    )

    # Para grafico
    ingresos_por_departamento = (
        registros
        .exclude(departamento_destino__isnull=True)
        .values(
            nombre=F('departamento_destino__nombre')
        )
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
    )

    # Por persona, por departamento, para gráfico
    ultimos_movimientos = (
        RegistroAcceso.objects
        .values('id_persona')
        .annotate(ultima_fecha=Max('fecha_hora'))
    )
    personas_dentro = RegistroAcceso.objects.filter(
        tipo_movimiento='INGRESO',
        id__in=[
            RegistroAcceso.objects.filter(
                id_persona=m['id_persona'],
                fecha_hora=m['ultima_fecha']
            ).values('id')[:1]
            for m in ultimos_movimientos
        ]
    )
    personas_dentro_por_departamento = (
        personas_dentro
        .exclude(departamento_destino__isnull=True)
        .values(
            nombre=F('departamento_destino__nombre')
        )
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
    )

    # ---------------------------------------
    # INGRESOS DE VEHÍCULOS (por período)
    # ---------------------------------------
    vehiculos_registros = registros.exclude(vehiculo__isnull=True)

    vehiculos_por_periodo = (
        vehiculos_registros
        .annotate(periodo=trunc)
        .values('periodo')
        .annotate(cantidad=Count('id'))
        .order_by('periodo')
    )

    for r in vehiculos_por_periodo:
        if isinstance(trunc, TruncDay):
            r['periodo'] = r['periodo'].strftime('%d/%m')
        elif isinstance(trunc, TruncWeek):
            inicio_semana = r['periodo']
            fin_semana = inicio_semana + timedelta(days=6)
            r['periodo'] = f"{inicio_semana.strftime('%d/%m')} - {fin_semana.strftime('%d/%m')}"
        else:
            r['periodo'] = r['periodo'].strftime('%b %Y')

    # ---------------------------------------
    # VEHÍCULOS DENTRO
    # ---------------------------------------
    ultimos_movimientos_vehiculos = (
        RegistroAcceso.objects
        .exclude(vehiculo__isnull=True)
        .values('vehiculo')
        .annotate(ultima_fecha=Max('fecha_hora'))
    )

    vehiculos_dentro = RegistroAcceso.objects.filter(
        tipo_movimiento='INGRESO',
        id__in=[
            RegistroAcceso.objects.filter(
                vehiculo=m['vehiculo'],
                fecha_hora=m['ultima_fecha']
            ).values('id')[:1]
            for m in ultimos_movimientos_vehiculos
        ]
    )

    vehiculos_dentro_por_marca = (
        vehiculos_dentro
        .values(nombre=F('vehiculo__marca__nombre'))
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')
    )

    context = {
        "tipos": tipos,
        "total_ingresos": total_ingresos,
        "promedio_dias_activos": promedio_dias_activos,
        "promedio_periodo": promedio_periodo,
        "dia_mas_ingresos": dia_mas_ingresos,
        "cantidad_mas_ingresos": cantidad_mas_ingresos_fmt,
        "hora_pico": hora_pico,
        "cantidad_hora_pico": cantidad_hora_pico,
        # "ingresos_por_dia": ingresos_por_dia,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "ingresos_por_periodo": list(ingresos_por_periodo),
        "ingresos_por_hora": ingresos_por_hora,
        "ingresos_por_tipo": list(ingresos_por_tipo),
        "detalle_accesos": detalle_accesos,
        "ingresos_por_departamento": list(ingresos_por_departamento),
        "personas_dentro_por_departamento": list(personas_dentro_por_departamento),
        "vehiculos_por_periodo": list(vehiculos_por_periodo),
        "vehiculos_dentro_por_marca": list(vehiculos_dentro_por_marca),
    }

    return render(request, "pages/estadisticos.html", context)


@solo_admin
@login_required
def reportes(request):
    # -------------------------
    # Datos base para filtros
    # -------------------------
    tipos = TipoEmpleado.objects.all()
    departamentos = Departamento.objects.all()

    # -------------------------
    # Leer filtros
    # -------------------------
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    tipo = request.GET.get("tipo")
    departamento = request.GET.get("departamento")
    tipo_reporte = request.GET.get("reporte", "accesos")

    formato = request.GET.get("formato")

    resultados = None
    columnas = []

    # -------------------------
    # Query base
    # -------------------------
    # Query base mejorado
    registros = RegistroAcceso.objects.select_related(
        "id_persona",
        "id_persona__id_tipo",
        "departamento_destino",
        "vehiculo",
        "vehiculo__marca",
        "vehiculo__modelo"
    )

    # -------------------------
    # Aplicar filtros
    # -------------------------
    if desde and hasta:
        registros = registros.filter(
            fecha_hora__date__range=[desde, hasta]
        )

    if tipo:
        registros = registros.filter(id_persona__id_tipo__nombre_tipo=tipo)

    if departamento:
        registros = registros.filter(departamento_destino__nombre=departamento)

    if formato == "pdf":
        return exportar_pdf(request, tipo_reporte=tipo_reporte)

    if formato == "excel":
        return exportar_excel(request, tipo_reporte=tipo_reporte)

    if formato == "csv":
        return exportar_csv(request, tipo_reporte=tipo_reporte)

    if tipo_reporte == "accesos":

        columnas = [
            "Fecha",
            "Persona",
            "Tipo",
            "Movimiento",
            "Departamento"
        ]

        resultados = [
            {
                "fecha": r.fecha_hora.strftime("%d/%m/%Y %I:%M %p"),
                "persona": f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                "tipo": r.id_persona.id_tipo.nombre_tipo,
                "movimiento": r.get_tipo_movimiento_display(),
                "departamento": r.departamento_destino.nombre if r.departamento_destino else None
            }
            for r in registros.order_by("-fecha_hora")
        ]



    elif tipo_reporte == "personas":

        columnas = [
            "Persona",
            "Tipo",
            "Total ingresos",
            "Último ingreso"
        ]

        resultados = [
            {
                "persona": f"{r['id_persona__nombres']} {r['id_persona__apellidos']}",
                "tipo": r['id_persona__id_tipo__nombre_tipo'],
                "total": r['total_ingresos'],
                "ultimo": r['ultimo_ingreso'].strftime("%d/%m/%Y %I:%M %p") if r['ultimo_ingreso'] else None
            }

            for r in (
                registros
                .filter(tipo_movimiento="INGRESO")
                .values(
                    "id_persona__nombres",
                    "id_persona__apellidos",
                    "id_persona__id_tipo__nombre_tipo"
                )

                .annotate(
                    total_ingresos=Count("id"),
                    ultimo_ingreso=Max("fecha_hora")
                )
            )
        ]


    elif tipo_reporte == "diario":

        columnas = [
            "Fecha",
            "Ingresos",
            "Salidas"
        ]

        resultados = (
            registros
            .values("fecha_hora__date")
            .annotate(
                ingresos=Count("id", filter=models.Q(tipo_movimiento="INGRESO")),
                salidas=Count("id", filter=models.Q(tipo_movimiento="EGRESO")),
            )
            .order_by("-fecha_hora__date")
        )

    elif tipo_reporte == "tipo":

        columnas = [
            "Tipo de persona",
            "Total ingresos"
        ]

        resultados = (
            registros
            .filter(tipo_movimiento="INGRESO")
            .values("id_persona__id_tipo__nombre_tipo")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

    elif tipo_reporte == "departamento":

        columnas = [
            "Departamento",
            "Total ingresos"
        ]

        resultados = (
            registros
            .filter(
                tipo_movimiento="INGRESO",
                departamento_destino__isnull=False
            )
            .values("departamento_destino__nombre")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        # ... (códigos anteriores de accesos, personas, diario, etc.)

    elif tipo_reporte == "vehiculos_detallado":
        columnas = [
            "Fecha",
            "Placa",
            "Vehículo",
            "Persona",
            "Movimiento",
            "Departamento"
        ]

        # Filtramos solo los registros que tienen vehículo
        registros_veh = registros.filter(vehiculo__isnull=False).order_by("-fecha_hora")

        resultados = [
            {
                "fecha": r.fecha_hora.strftime("%d/%m/%Y %I:%M %p"),
                "placa": r.vehiculo.placa,
                "vehiculo": f"{r.vehiculo.marca.nombre} {r.vehiculo.modelo.nombre}",
                "persona": f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                "movimiento": r.get_tipo_movimiento_display(),
                "departamento": r.departamento_destino.nombre if r.departamento_destino else "—"
            }
            for r in registros_veh
        ]

    elif tipo_reporte == "vehiculos_resumen":
        columnas = [
            "Placa",
            "Marca/Modelo",
            "Total Movimientos",
            "Último Acceso"
        ]

        # Agrupamos por vehículo
        resultados = [
            {
                "placa": r['vehiculo__placa'],
                "vehiculo": f"{r['vehiculo__marca__nombre']} {r['vehiculo__modelo__nombre']}",
                "total": r['total'],
                "ultimo": r['ultimo'].strftime("%d/%m/%Y %I:%M %p") if r['ultimo'] else "—"
            }
            for r in (
                registros
                .filter(vehiculo__isnull=False)
                .values(
                    "vehiculo__placa",
                    "vehiculo__marca__nombre",
                    "vehiculo__modelo__nombre"
                )
                .annotate(
                    total=Count("id"),
                    ultimo=Max("fecha_hora")
                )
                .order_by("-total")
            )
        ]

    paginator = Paginator(resultados, 20)  # 10 filas por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "tipos": tipos,
        "departamentos": departamentos,
        "resultados": resultados,
        "columnas": columnas,
        "filtros": request.GET,

        "page_obj": page_obj,
        "request": request,
    }

    return render(request, "pages/reportes.html", context)


@solo_admin
@login_required
def perfil(request):
    persona = request.user.persona

    dict = {"persona": persona}

    return render(request, "pages/perfil.html", dict)


def exportar_historial_pdf(request, persona_id):
    persona = Persona.objects.get(id=persona_id)
    registros = RegistroAcceso.objects.filter(id_persona=persona).order_by('fecha_hora')

    # Crear respuesta HTTP con tipo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historial_{persona.nombres}.pdf"'

    # Crear PDF
    c = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, f"Historial de accesos: {persona.nombres} {persona.apellidos}")
    y -= 30

    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Fecha y hora")
    c.drawString(200, y, "Tipo movimiento")
    c.drawString(350, y, "Departamento")
    y -= 20

    c.setFont("Helvetica", 12)
    for reg in registros:
        if y < 50:  # nueva página
            c.showPage()
            y = height - 50

        # c.drawString(50, y, reg.fecha_hora.strftime("%d/%m/%Y %I:%M %p"))
        fecha_local = timezone.localtime(reg.fecha_hora)

        c.drawString(
            50,
            y,
            fecha_local.strftime("%d/%m/%Y %I:%M %p")
        )

        c.drawString(200, y, reg.tipo_movimiento)
        depto = reg.departamento_destino.nombre if reg.departamento_destino else "-"
        c.drawString(350, y, depto)
        y -= 20

    c.save()
    return response

def exportar_historial_pdf_vehiculos(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    registros = RegistroAcceso.objects.filter(vehiculo=vehiculo).order_by('fecha_hora')

    # Crear respuesta HTTP con tipo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historial_{vehiculo.placa}.pdf"'

    # Crear PDF
    c = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, f"Historial de accesos: Vehículo con placa {vehiculo.placa}")
    y -= 30

    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Fecha y hora")
    c.drawString(200, y, "Tipo movimiento")
    c.drawString(350, y, "Departamento")
    y -= 20

    c.setFont("Helvetica", 12)
    for reg in registros:
        if y < 50:  # nueva página
            c.showPage()
            y = height - 50

        # c.drawString(50, y, reg.fecha_hora.strftime("%d/%m/%Y %I:%M %p"))
        fecha_local = timezone.localtime(reg.fecha_hora)

        c.drawString(
            50,
            y,
            fecha_local.strftime("%d/%m/%Y %I:%M %p")
        )

        c.drawString(200, y, reg.tipo_movimiento)
        depto = reg.departamento_destino.nombre if reg.departamento_destino else "-"
        c.drawString(350, y, depto)
        y -= 20

    c.save()
    return response


def exportar_historial_excel(request, persona_id):
    persona = Persona.objects.get(id=persona_id)
    registros = RegistroAcceso.objects.filter(id_persona=persona).order_by('fecha_hora')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Encabezados
    headers = ["Fecha y hora", "Tipo movimiento", "Departamento"]
    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # Datos
    for row_num, reg in enumerate(registros, 2):
        ws[f"A{row_num}"] = reg.fecha_hora.strftime("%d/%m/%Y %I:%M %p")
        ws[f"B{row_num}"] = reg.tipo_movimiento
        ws[f"C{row_num}"] = reg.departamento_destino.nombre if reg.departamento_destino else "-"

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=historial_{persona.nombres}.xlsx'
    wb.save(response)
    return response

def exportar_historial_excel_vehiculos(request, vehiculo_id):
    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
    registros = RegistroAcceso.objects.filter(vehiculo=vehiculo).order_by('fecha_hora')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Encabezados
    headers = ["Fecha y hora", "Tipo movimiento", "Departamento"]
    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # Datos
    for row_num, reg in enumerate(registros, 2):
        fecha_local = timezone.localtime(reg.fecha_hora)
        # ws[f"A{row_num}"] = reg.fecha_hora.strftime("%d/%m/%Y %I:%M %p")
        ws[f"A{row_num}"] = fecha_local.strftime("%d/%m/%Y %I:%M %p")
        ws[f"B{row_num}"] = reg.tipo_movimiento
        ws[f"C{row_num}"] = reg.departamento_destino.nombre if reg.departamento_destino else "-"

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=historial_{vehiculo.placa}.xlsx'
    wb.save(response)
    return response


def obtener_queryset_reporte(request):
    qs = RegistroAcceso.objects.select_related(
        "id_persona",
        "departamento_destino",
    )

    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    tipo = request.GET.get("tipo")
    departamento = request.GET.get("departamento")

    if desde:
        qs = qs.filter(fecha_hora__date__gte=desde)

    if hasta:
        qs = qs.filter(fecha_hora__date__lte=hasta)

    if tipo:
        qs = qs.filter(id_persona__id_tipo__nombre_tipo=tipo)

    if departamento:
        qs = qs.filter(departamento_destino__id=departamento)

    return qs.order_by("-fecha_hora")


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors


def exportar_pdf(request, tipo_reporte="accesos"):
    queryset = obtener_queryset_reporte(request)

    response = HttpResponse(content_type="application/pdf")
    filename = f"reporte_{tipo_reporte}.pdf"
    response["Content-Disposition"] = f"attachment; filename={filename}"

    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []

    # Variables para filtros (usadas en reportes con agregación manual)
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    tipo = request.GET.get("tipo")
    departamento = request.GET.get("departamento")

    # ---------------------------
    # Lógica de datos por reporte
    # ---------------------------
    if tipo_reporte == "accesos":
        columnas = ["Fecha y hora", "Persona", "Tipo", "Movimiento", "Departamento"]
        data = [columnas]
        for r in queryset:
            data.append([
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.id_persona.id_tipo.nombre_tipo,
                r.get_tipo_movimiento_display(),
                r.departamento_destino.nombre if r.departamento_destino else "-"
            ])

    elif tipo_reporte == "vehiculos_detallado":
        columnas = ["Fecha", "Placa", "Vehículo", "Persona", "Movimiento"]
        data = [columnas]
        # Filtrar solo registros que tengan vehículo asociado
        qs = queryset.filter(vehiculo__isnull=False).select_related('vehiculo__marca', 'vehiculo__modelo')
        for r in qs:
            data.append([
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                r.vehiculo.placa,
                f"{r.vehiculo.marca.nombre} {r.vehiculo.modelo.nombre}",
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.get_tipo_movimiento_display()
            ])

    elif tipo_reporte == "vehiculos_resumen":
        columnas = ["Placa", "Marca/Modelo", "Total Movimientos", "Último Acceso"]
        data = [columnas]
        qs = queryset.filter(vehiculo__isnull=False).values(
            "vehiculo__placa",
            "vehiculo__marca__nombre",
            "vehiculo__modelo__nombre"
        ).annotate(
            total=Count("id"),
            ultimo=Max("fecha_hora")
        ).order_by("-total")

        for r in qs:
            data.append([
                r['vehiculo__placa'],
                f"{r['vehiculo__marca__nombre']} {r['vehiculo__modelo__nombre']}",
                r['total'],
                r['ultimo'].strftime("%d/%m/%Y %H:%M") if r['ultimo'] else "-"
            ])

    elif tipo_reporte == "personas":
        columnas = ["Persona", "Tipo", "Total ingresos", "Último ingreso"]
        data = [columnas]
        qs = RegistroAcceso.objects.filter(tipo_movimiento="INGRESO")

        if desde: qs = qs.filter(fecha_hora__date__gte=desde)
        if hasta: qs = qs.filter(fecha_hora__date__lte=hasta)
        if tipo: qs = qs.filter(id_persona__id_tipo__nombre_tipo=tipo)
        if departamento: qs = qs.filter(departamento_destino__nombre=departamento)

        qs = qs.values(
            "id_persona__nombres",
            "id_persona__apellidos",
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(
            total_ingresos=Count("id"),
            ultimo_ingreso=Max("fecha_hora")
        ).order_by("id_persona__apellidos")

        for r in qs:
            data.append([
                f"{r['id_persona__nombres']} {r['id_persona__apellidos']}",
                r['id_persona__id_tipo__nombre_tipo'],
                r['total_ingresos'],
                r['ultimo_ingreso'].strftime("%d/%m/%Y %H:%M") if r['ultimo_ingreso'] else "-"
            ])

    elif tipo_reporte == "diario":
        columnas = ["Fecha", "Ingresos", "Salidas"]
        data = [columnas]
        qs = queryset.values("fecha_hora__date").annotate(
            ingresos=Count("id", filter=models.Q(tipo_movimiento="INGRESO")),
            salidas=Count("id", filter=models.Q(tipo_movimiento="EGRESO"))
        ).order_by("fecha_hora__date")

        for r in qs:
            data.append([
                r['fecha_hora__date'].strftime("%d/%m/%Y"),
                r['ingresos'],
                r['salidas']
            ])

    elif tipo_reporte == "tipo":
        columnas = ["Tipo de persona", "Total ingresos"]
        data = [columnas]
        qs = queryset.filter(tipo_movimiento="INGRESO").values(
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(total=Count("id")).order_by("-total")

        for r in qs:
            data.append([
                r['id_persona__id_tipo__nombre_tipo'],
                r['total']
            ])

    elif tipo_reporte == "departamento":
        columnas = ["Departamento", "Total ingresos"]
        data = [columnas]
        qs = queryset.filter(
            tipo_movimiento="INGRESO",
            departamento_destino__isnull=False
        ).values(
            "departamento_destino__nombre"
        ).annotate(total=Count("id")).order_by("-total")

        for r in qs:
            data.append([
                r['departamento_destino__nombre'],
                r['total']
            ])

    # ---------------------------
    # Generación de la tabla PDF
    # ---------------------------
    colWidths = []
    for col_idx in range(len(data[0])):
        max_width = max(stringWidth(str(row[col_idx]), "Helvetica", 10) for row in data)
        colWidths.append(max_width + 15)  # Un poco más de margen

    tabla = Table(data, colWidths=colWidths)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response


def exportar_excel(request, tipo_reporte="accesos"):
    queryset = obtener_queryset_reporte(request)

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    filas = []
    columnas = []

    # ------------------------
    # Generar datos según reporte
    # ------------------------
    if tipo_reporte == "accesos":
        columnas = ["Fecha y hora", "Persona", "Tipo", "Movimiento", "Departamento"]
        filas = [
            [
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.id_persona.id_tipo.nombre_tipo,
                r.get_tipo_movimiento_display(),
                r.departamento_destino.nombre if r.departamento_destino else "-"
            ]
            for r in queryset
        ]

    elif tipo_reporte == "vehiculos_detallado":
        columnas = ["Fecha y hora", "Placa", "Vehículo", "Persona", "Movimiento"]
        qs = queryset.filter(vehiculo__isnull=False).select_related('vehiculo__marca', 'vehiculo__modelo')
        filas = [
            [
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                r.vehiculo.placa,
                f"{r.vehiculo.marca.nombre} {r.vehiculo.modelo.nombre}",
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.get_tipo_movimiento_display()
            ]
            for r in qs
        ]

    elif tipo_reporte == "vehiculos_resumen":
        columnas = ["Placa", "Marca/Modelo", "Total Movimientos", "Último Acceso"]
        qs = queryset.filter(vehiculo__isnull=False).values(
            "vehiculo__placa",
            "vehiculo__marca__nombre",
            "vehiculo__modelo__nombre"
        ).annotate(
            total=Count("id"),
            ultimo=Max("fecha_hora")
        ).order_by("-total")

        filas = [
            [
                r['vehiculo__placa'],
                f"{r['vehiculo__marca__nombre']} {r['vehiculo__modelo__nombre']}",
                r['total'],
                r['ultimo'].strftime("%d/%m/%Y %H:%M") if r['ultimo'] else "-"
            ]
            for r in qs
        ]

    elif tipo_reporte == "personas":
        columnas = ["Persona", "Tipo", "Total ingresos", "Último ingreso"]
        qs = queryset.filter(tipo_movimiento="INGRESO").values(
            "id_persona__nombres",
            "id_persona__apellidos",
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(
            total_ingresos=Count("id"),
            ultimo_ingreso=Max("fecha_hora")
        ).order_by("id_persona__nombres")

        filas = [
            [
                f"{r['id_persona__nombres']} {r['id_persona__apellidos']}",
                r['id_persona__id_tipo__nombre_tipo'],
                r['total_ingresos'],
                r['ultimo_ingreso'].strftime("%d/%m/%Y %H:%M") if r['ultimo_ingreso'] else "-"
            ]
            for r in qs
        ]

    elif tipo_reporte == "diario":
        columnas = ["Fecha", "Ingresos", "Salidas"]
        qs = queryset.values("fecha_hora__date").annotate(
            ingresos=Count("id", filter=models.Q(tipo_movimiento="INGRESO")),
            salidas=Count("id", filter=models.Q(tipo_movimiento="EGRESO"))
        ).order_by("fecha_hora__date")
        filas = [
            [
                r['fecha_hora__date'].strftime("%d/%m/%Y"),
                r['ingresos'],
                r['salidas']
            ]
            for r in qs
        ]

    elif tipo_reporte == "tipo":
        columnas = ["Tipo de persona", "Total ingresos"]
        qs = queryset.filter(tipo_movimiento="INGRESO").values(
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(total=Count("id")).order_by("-total")
        filas = [
            [r['id_persona__id_tipo__nombre_tipo'], r['total']]
            for r in qs
        ]

    elif tipo_reporte == "departamento":
        columnas = ["Departamento", "Total ingresos"]
        qs = queryset.filter(tipo_movimiento="INGRESO", departamento_destino__isnull=False).values(
            "departamento_destino__nombre"
        ).annotate(total=Count("id")).order_by("-total")
        filas = [
            [r['departamento_destino__nombre'], r['total']]
            for r in qs
        ]

    # ------------------------
    # Escribir encabezados
    # ------------------------
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ------------------------
    # Escribir filas
    # ------------------------
    for row_num, row_data in enumerate(filas, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")

    # ------------------------
    # Ajustar ancho columnas
    # ------------------------
    if filas:
        for i, col in enumerate(columnas, 1):
            max_length = max(
                [len(str(row[i - 1])) for row in filas] + [len(col)]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # ------------------------
    # Devolver archivo
    # ------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"reporte_{tipo_reporte}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


def exportar_csv(request, tipo_reporte="accesos"):
    queryset = obtener_queryset_reporte(request)

    response = HttpResponse(
        content_type="text/csv; charset=utf-8"
    )
    # Nombre de archivo dinámico
    filename = f"reporte_{tipo_reporte}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    response.write("\ufeff")  # BOM para compatibilidad con Excel

    writer = csv.writer(response)

    columnas = []
    filas = []

    # ------------------------
    # ACCESOS DETALLADOS
    # ------------------------
    if tipo_reporte == "accesos":
        columnas = ["Fecha y hora", "Persona", "Tipo", "Movimiento", "Departamento"]
        filas = [
            [
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.id_persona.id_tipo.nombre_tipo,
                r.get_tipo_movimiento_display(),
                r.departamento_destino.nombre if r.departamento_destino else "-"
            ]
            for r in queryset
        ]

    # ------------------------
    # ACCESOS VEHÍCULOS DETALLADO
    # ------------------------
    elif tipo_reporte == "vehiculos_detallado":
        columnas = ["Fecha y hora", "Placa", "Vehículo", "Persona", "Movimiento"]
        qs = queryset.filter(vehiculo__isnull=False).select_related('vehiculo__marca', 'vehiculo__modelo')
        filas = [
            [
                r.fecha_hora.strftime("%d/%m/%Y %H:%M"),
                r.vehiculo.placa,
                f"{r.vehiculo.marca.nombre} {r.vehiculo.modelo.nombre}",
                f"{r.id_persona.nombres} {r.id_persona.apellidos}",
                r.get_tipo_movimiento_display()
            ]
            for r in qs
        ]

    # ------------------------
    # RESUMEN POR VEHÍCULO
    # ------------------------
    elif tipo_reporte == "vehiculos_resumen":
        columnas = ["Placa", "Marca/Modelo", "Total Movimientos", "Último Acceso"]
        qs = queryset.filter(vehiculo__isnull=False).values(
            "vehiculo__placa",
            "vehiculo__marca__nombre",
            "vehiculo__modelo__nombre"
        ).annotate(
            total=Count("id"),
            ultimo=Max("fecha_hora")
        ).order_by("-total")

        filas = [
            [
                r['vehiculo__placa'],
                f"{r['vehiculo__marca__nombre']} {r['vehiculo__modelo__nombre']}",
                r['total'],
                r['ultimo'].strftime("%d/%m/%Y %H:%M") if r['ultimo'] else "-"
            ]
            for r in qs
        ]

    # ------------------------
    # RESUMEN POR PERSONA
    # ------------------------
    elif tipo_reporte == "personas":
        columnas = ["Persona", "Tipo", "Total ingresos", "Último ingreso"]

        qs = queryset.filter(tipo_movimiento="INGRESO").values(
            "id_persona__nombres",
            "id_persona__apellidos",
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(
            total_ingresos=Count("id"),
            ultimo_ingreso=Max("fecha_hora")
        ).order_by("id_persona__apellidos")

        filas = [
            [
                f"{r['id_persona__nombres']} {r['id_persona__apellidos']}",
                r['id_persona__id_tipo__nombre_tipo'],
                r['total_ingresos'],
                r['ultimo_ingreso'].strftime("%d/%m/%Y %H:%M") if r['ultimo_ingreso'] else "-"
            ]
            for r in qs
        ]

    # ------------------------
    # RESUMEN DIARIO
    # ------------------------
    elif tipo_reporte == "diario":
        columnas = ["Fecha", "Ingresos", "Salidas"]

        qs = queryset.values("fecha_hora__date").annotate(
            ingresos=Count("id", filter=models.Q(tipo_movimiento="INGRESO")),
            salidas=Count("id", filter=models.Q(tipo_movimiento="EGRESO"))
        ).order_by("fecha_hora__date")

        filas = [
            [
                r["fecha_hora__date"].strftime("%d/%m/%Y"),
                r["ingresos"],
                r["salidas"]
            ]
            for r in qs
        ]

    # ------------------------
    # POR TIPO DE PERSONA
    # ------------------------
    elif tipo_reporte == "tipo":
        columnas = ["Tipo de persona", "Total ingresos"]

        qs = queryset.filter(tipo_movimiento="INGRESO").values(
            "id_persona__id_tipo__nombre_tipo"
        ).annotate(total=Count("id")).order_by("-total")

        filas = [
            [r["id_persona__id_tipo__nombre_tipo"], r["total"]]
            for r in qs
        ]

    # ------------------------
    # POR DEPARTAMENTO
    # ------------------------
    elif tipo_reporte == "departamento":
        columnas = ["Departamento", "Total ingresos"]

        qs = queryset.filter(
            tipo_movimiento="INGRESO",
            departamento_destino__isnull=False
        ).values(
            "departamento_destino__nombre"
        ).annotate(total=Count("id")).order_by("-total")

        filas = [
            [r["departamento_destino__nombre"], r["total"]]
            for r in qs
        ]

    # ------------------------
    # Escribir CSV
    # ------------------------
    writer.writerow(columnas)
    writer.writerows(filas)

    return response


def recuperar_contraseña(request):
    dict = {}

    if request.POST:

        usuario = Usuario.objects.filter(email=request.POST["email"]).first()

        if usuario:

            password_temporal = secrets.token_urlsafe(8)
            usuario.set_password(password_temporal)
            usuario.debe_cambiar_password = True
            usuario.save()

            html_content = f"""
                            Hola {usuario.first_name},
    
                            Se ha cambiar su dirección de correo electrónico.
    
                            Usuario: {usuario.email}
                            Contraseña temporal: {password_temporal}
    
                            Por seguridad deberás cambiar esta contraseña al iniciar sesión.
                            """

            enviar_correo(
                usuario.email,
                "Acceso al sistema",
                html_content
            )

            # send_mail(
            #     'Acceso al sistema',
            #     f'''
            #                 Hola {usuario.first_name},
            #
            #                 Se ha cambiar su dirección de correo electrónico.
            #
            #                 Usuario: {usuario.email}
            #                 Contraseña temporal: {password_temporal}
            #
            #                 Por seguridad deberás cambiar esta contraseña al iniciar sesión.
            #                 ''',
            #     'no-reply@sistema.com',
            #     [request.POST['email']],
            #     fail_silently=False,
            # )
            return redirect('modelNewApp:home')
        else:
            messages.error(request, "No existe un usuario con ese correo.")
            return render(
                request,
                "pages/recuperar_contraseña.html",
                {
                    "error": "No existe un usuario con ese correo.",
                })

    return render(request, "pages/recuperar_contraseña.html", dict)

# ---------------------------------------------------------------------------------------
#     Respaldo simple de la base de datos
# ---------------------------------------------------------------------------------------

import os
import shutil
from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib.auth.decorators import login_required
# Asumo que solo_admin es tu decorador personalizado
from .decorators import solo_admin

DB_FILE = os.path.join(settings.BASE_DIR, 'db.sqlite3')


@login_required
@solo_admin
def db_panel(request):
    return render(request, 'pages/admin_db_panel.html')


@login_required
@solo_admin
def db_download(request):
    if os.path.exists(DB_FILE):
        import datetime
        fecha = datetime.datetime.now().strftime("%Y_%m_%d")
        return FileResponse(
            open(DB_FILE, 'rb'),
            as_attachment=True,
            filename=f"respaldo_asistencia_{fecha}.sqlite3"
        )
    return HttpResponse("Base de datos no encontrada", status=404)


@login_required
@solo_admin
def db_restore(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get('db_file')

        if uploaded_file and uploaded_file.name.endswith('.sqlite3'):
            from django.db import connection
            connection.close()

            # Recomendación para la tesis: Crear un backup temporal antes de sobrescribir
            # Esto se llama "Fail-safe mechanism"
            backup_temp = DB_FILE + ".tmp"
            shutil.copy2(DB_FILE, backup_temp)

            try:
                with open(DB_FILE, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)

                # Si todo sale bien, borramos el temporal
                if os.path.exists(backup_temp):
                    os.remove(backup_temp)

            except Exception as e:
                # Si algo falla durante la escritura, restauramos el original
                shutil.move(backup_temp, DB_FILE)
                return HttpResponse(f"Error crítico en la restauración: {e}", status=500)

            return redirect('modelNewApp:db_panel')

    return HttpResponse("Método no permitido", status=405)